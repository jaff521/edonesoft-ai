#!/usr/bin/env python3
"""
通用 Excel / CSV 文件读取脚本
用于从客户上传的 Excel (.xlsx) 或 CSV (.csv) 文件中提取表格数据，
输出标准结构化 JSON 供 LLM 进行开票明细或其它业务数据的字段映射。

用法:
  python3 excel_reader.py "/path/to/file.xlsx"
  python3 excel_reader.py "/path/to/file.csv"
  python3 excel_reader.py "/path/to/file.xlsx" --sheet "Sheet1" --max-rows 100
"""

import sys
import os
import csv
import json
import argparse
from typing import Dict, Any, List, Optional


def detect_file_encoding(file_path: str) -> str:
    """尝试检测 CSV 文件的编码，依次尝试 utf-8-sig, utf-8, gbk, gb18030, utf-16。"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030", "utf-16"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8-sig"  # 默认回退


def read_csv_file(file_path: str, max_rows: int = 10000) -> Dict[str, Any]:
    """读取 CSV 文件并转换为标准结构化 JSON。"""
    encoding = detect_file_encoding(file_path)

    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        # 尝试嗅探分隔符
        sample = f.read(4096)
        f.seek(0)
        delimiter = ","
        if sample:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","

        reader = csv.reader(f, delimiter=delimiter)
        raw_rows = []
        for row in reader:
            # 过滤全空行
            if any(str(cell).strip() for cell in row):
                raw_rows.append([str(cell).strip() for cell in row])

    if not raw_rows:
        return {
            "success": False,
            "message": "CSV 文件为空或未包含有效数据行",
            "sheets": []
        }

    # 寻找表头行（第一行有 2 个以上非空列的行）
    header_idx = 0
    for i, row in enumerate(raw_rows[:10]):
        non_empty = [c for c in row if c]
        if len(non_empty) >= 2:
            header_idx = i
            break

    headers = raw_rows[header_idx]
    # 清理表头列名，去重及填充空列名
    clean_headers = []
    seen = {}
    for col_i, h in enumerate(headers):
        h_name = h if h else f"列_{col_i + 1}"
        if h_name in seen:
            seen[h_name] += 1
            h_name = f"{h_name}_{seen[h_name]}"
        else:
            seen[h_name] = 1
        clean_headers.append(h_name)

    data_rows = []
    truncated = False
    for row in raw_rows[header_idx + 1:]:
        if max_rows > 0 and len(data_rows) >= max_rows:
            truncated = True
            break
        row_dict = {}
        for col_i, h_name in enumerate(clean_headers):
            val = row[col_i] if col_i < len(row) else ""
            row_dict[h_name] = val
        data_rows.append(row_dict)

    msg = f"成功读取 CSV 文件，共 {len(data_rows)} 行数据"
    if truncated:
        msg += f"（已按限制截断前 {max_rows} 行）"

    return {
        "success": True,
        "file": os.path.basename(file_path),
        "sheets": [
          {
              "name": "CSV_DATA",
              "headers": clean_headers,
              "row_count": len(data_rows),
              "rows": data_rows
          }
        ],
        "total_rows": len(data_rows),
        "message": msg
    }


def read_excel_file(file_path: str, target_sheet: Optional[str] = None, max_rows: int = 10000) -> Dict[str, Any]:
    """使用 openpyxl 读取 .xlsx 文件并转换为标准结构化 JSON。"""
    try:
        import openpyxl
    except ImportError:
        return {
            "success": False,
            "message": "系统未安装 openpyxl 库，请先执行 pip install openpyxl",
            "sheets": []
        }

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {
            "success": False,
            "message": f"无法打开 Excel 文件: {str(e)}",
            "sheets": []
        }

    sheet_names = wb.sheetnames
    if not sheet_names:
        return {
            "success": False,
            "message": "Excel 文件不包含任何工作表(Sheet)",
            "sheets": []
        }

    selected_sheets = sheet_names
    if target_sheet:
        if target_sheet in sheet_names:
            selected_sheets = [target_sheet]
        else:
            return {
                "success": False,
                "message": f"未找到指定的工作表 '{target_sheet}'，可选工作表: {sheet_names}",
                "sheets": []
            }

    result_sheets = []
    grand_total_rows = 0

    for s_name in selected_sheets:
        ws = wb[s_name]

        # 提取合并单元格的取值字典
        merged_cell_map = {}
        for mrange in ws.merged_cells.ranges:
            top_left_val = ws.cell(row=mrange.min_row, column=mrange.min_col).value
            for row_idx in range(mrange.min_row, mrange.max_row + 1):
                for col_idx in range(mrange.min_col, mrange.max_col + 1):
                    merged_cell_map[(row_idx, col_idx)] = top_left_val

        # 逐行读取原始单元格值
        raw_rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            row_vals = []
            for c_idx, cell in enumerate(row, start=1):
                if (r_idx, c_idx) in merged_cell_map:
                    val = merged_cell_map[(r_idx, c_idx)]
                else:
                    val = cell.value

                # 格式化不同类型的数据
                if val is None:
                    val_str = ""
                elif hasattr(val, "strftime"):
                    val_str = val.strftime("%Y-%m-%d")
                elif isinstance(val, float):
                    # 如果能转为整数则简化
                    val_str = int(val) if val.is_integer() else val
                else:
                    val_str = str(val).strip()

                row_vals.append(val_str)

            # 过滤全空行
            if any(str(v).strip() for v in row_vals):
                raw_rows.append(row_vals)

        if not raw_rows:
            continue

        # 寻找表头行
        header_idx = 0
        for i, r in enumerate(raw_rows[:10]):
            non_empty = [c for c in r if str(c).strip()]
            if len(non_empty) >= 2:
                header_idx = i
                break

        headers = raw_rows[header_idx]
        clean_headers = []
        seen = {}
        for col_i, h in enumerate(headers):
            h_str = str(h).strip() if h is not None else ""
            h_name = h_str if h_str else f"列_{col_i + 1}"
            if h_name in seen:
                seen[h_name] += 1
                h_name = f"{h_name}_{seen[h_name]}"
            else:
                seen[h_name] = 1
            clean_headers.append(h_name)

        data_rows = []
        for r in raw_rows[header_idx + 1:]:
            if max_rows > 0 and len(data_rows) >= max_rows:
                break
            row_dict = {}
            for col_i, h_name in enumerate(clean_headers):
                val = r[col_i] if col_i < len(r) else ""
                row_dict[h_name] = val
            data_rows.append(row_dict)

        result_sheets.append({
            "name": s_name,
            "headers": clean_headers,
            "row_count": len(data_rows),
            "rows": data_rows
        })
        grand_total_rows += len(data_rows)

    return {
        "success": True,
        "file": os.path.basename(file_path),
        "sheets": result_sheets,
        "total_rows": grand_total_rows,
        "message": f"成功读取 {len(result_sheets)} 个工作表，共 {grand_total_rows} 行数据"
    }


def main():
    parser = argparse.ArgumentParser(description="通用 Excel / CSV 文件读取工具")
    parser.add_argument("file_path", help="要读取的 Excel (.xlsx) 或 CSV (.csv) 文件路径")
    parser.add_argument("--sheet", help="指定要读取的工作表(Sheet)名称", default=None)
    parser.add_argument("--max-rows", type=int, help="最多读取的行数 (默认 10000，设置 0 表示不限制)", default=10000)

    args = parser.parse_args()

    file_path = args.file_path.strip('"').strip("'")
    if not os.path.exists(file_path):
        res = {
            "success": False,
            "message": f"文件不存在: {file_path}",
            "sheets": []
        }
        print(json.dumps(res, ensure_ascii=False, indent=2))
        return

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        res = read_csv_file(file_path, max_rows=args.max_rows)
    elif ext in [".xlsx", ".xlsm"]:
        res = read_excel_file(file_path, target_sheet=args.sheet, max_rows=args.max_rows)
    elif ext == ".xls":
        # 如果是旧版 .xls 格式，尝试通过 xlrd 或 pandas 读取，否则给友情提示
        try:
            import pandas as pd
            df = pd.read_excel(file_path, sheet_name=args.sheet or 0)
            headers = [str(c).strip() for c in df.columns]
            rows = df.fillna("").to_dict(orient="records")
            clean_rows = []
            for idx, r in enumerate(rows):
                if args.max_rows > 0 and idx >= args.max_rows:
                    break
                clean_r = {str(k).strip(): (int(v) if isinstance(v, float) and v.is_integer() else str(v).strip()) for k, v in r.items()}
                clean_rows.append(clean_r)

            res = {
                "success": True,
                "file": os.path.basename(file_path),
                "sheets": [{
                    "name": args.sheet or "Sheet1",
                    "headers": headers,
                    "row_count": len(clean_rows),
                    "rows": clean_rows
                }],
                "total_rows": len(clean_rows),
                "message": f"成功读取 .xls 文件，共 {len(clean_rows)} 行数据"
            }
        except Exception as e:
            res = {
                "success": False,
                "message": f"旧版 .xls 文件读取失败 ({str(e)})，请提示客户将其另存为 .xlsx 格式后重新发送",
                "sheets": []
            }
    else:
        res = {
            "success": False,
            "message": f"不支持的文件后缀 '{ext}'，目前仅支持 .xlsx, .csv 格式",
            "sheets": []
        }

    print(json.dumps(res, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    main()
